# ============================================================
# main.py
# Digital Infrastructure Intelligence Platform
# ============================================================

from fastapi import FastAPI, Request, BackgroundTasks
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from fastapi.responses import JSONResponse, RedirectResponse, StreamingResponse
from contextlib import asynccontextmanager
import asyncio
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
import csv

from database import db
from scanner import NetworkScanner
from monitor import AlertMonitor
from auth import auth_manager
from email_service import email_service


# ============================================================
# LIFESPAN — pornire si oprire aplicatie
# ============================================================
@asynccontextmanager
async def lifespan(app: FastAPI):
    print("\n" + "="*50)
    print(" Digital Infrastructure Intelligence Platform")
    print("="*50)

    test = db.test_connection()
    if test["status"] == "ok":
        print(f"✓ DB conectat: {test['server_name']}")
    else:
        print(f"✗ DB eroare: {test['message']}")

    monitor = AlertMonitor(check_interval_seconds=300)
    task    = asyncio.create_task(monitor.run())
    print("✓ Monitor pornit (verificare la fiecare 5 minute)")
    print("✓ Aplicatie disponibila la: http://127.0.0.1:8080")
    print("="*50 + "\n")

    yield

    task.cancel()
    print("\n[APP] Aplicatie oprita.")


# ============================================================
# APP INSTANCE
# ============================================================
app = FastAPI(
    title       = "Digital Infrastructure Intelligence Platform",
    description = "Network discovery, inventory and monitoring",
    version     = "1.0.0",
    lifespan    = lifespan
)

app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")
scanner   = NetworkScanner()


# ============================================================
# RUTE AUTENTIFICARE
# ============================================================

@app.get("/login")
async def page_login(request: Request):
    """Pagina de login — singura pagina publica."""
    user = auth_manager.get_current_user(request)
    if user:
        return RedirectResponse(url="/", status_code=302)
    return templates.TemplateResponse(request=request, name="login.html")


@app.post("/api/auth/login")
async def api_login(request: Request):
    """Primeste username + parola, verifica si creeaza sesiune."""
    try:
        data       = await request.json()
        username   = data.get("username", "").strip()
        password   = data.get("password", "")
        ip_address = request.client.host

        result = auth_manager.login(username, password, ip_address)

        if result["success"]:
            response = JSONResponse(content={
                "success": True,
                "user":    result["user"]
            })
            response.set_cookie(
                key      = "session_token",
                value    = result["token"],
                httponly = True,
                max_age  = 8 * 3600,
                samesite = "lax"
            )
            return response
        else:
            return JSONResponse(
                content={"success": False, "message": result["message"]},
                status_code=401
            )
    except Exception as e:
        return JSONResponse(
            content={"success": False, "message": str(e)},
            status_code=500
        )


@app.post("/api/auth/logout")
async def api_logout(request: Request):
    """Invalideaza sesiunea si sterge cookie-ul."""
    auth_manager.logout(request)
    response = JSONResponse(content={"success": True})
    response.delete_cookie("session_token")
    return response


@app.get("/api/auth/me")
async def api_me(request: Request):
    """Returneaza informatiile userului logat."""
    user = auth_manager.get_current_user(request)
    if not user:
        return JSONResponse(content={"authenticated": False}, status_code=401)
    return JSONResponse(content={"authenticated": True, "user": user})


# ============================================================
# RUTE PAGINI (protejate — necesita login)
# ============================================================

@app.get("/")
async def page_dashboard(request: Request):
    user = auth_manager.get_current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=302)
    return templates.TemplateResponse(
        request=request, name="dashboard.html", context={"user": user}
    )


@app.get("/inventory")
async def page_inventory(request: Request):
    user = auth_manager.get_current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=302)
    return templates.TemplateResponse(
        request=request, name="inventory.html", context={"user": user}
    )


@app.get("/scan")
async def page_scan(request: Request):
    user = auth_manager.get_current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=302)
    return templates.TemplateResponse(
        request=request, name="scan.html", context={"user": user}
    )


@app.get("/alerts")
async def page_alerts(request: Request):
    user = auth_manager.get_current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=302)
    return templates.TemplateResponse(
        request=request, name="alerts.html", context={"user": user}
    )


@app.get("/audit")
async def page_audit(request: Request):
    user = auth_manager.get_current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=302)
    return templates.TemplateResponse(
        request=request, name="audit.html", context={"user": user}
    )


@app.get("/device/{ip_address}")
async def page_device_detail(request: Request, ip_address: str):
    user = auth_manager.get_current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    # Logam vizualizarea
    auth_manager.log_action(
        action      = "VIEW_DEVICE",
        user_id     = user["user_id"],
        username    = user["username"],
        entity_type = "device",
        details     = f"Vizualizare device: {ip_address.replace('-', '.')}",
        ip_address  = request.client.host
    )

    return templates.TemplateResponse(
        request=request,
        name="device_detail.html",
        context={
            "ip_address": ip_address.replace("-", "."),
            "user": user
        }
    )


# ============================================================
# API — DASHBOARD
# ============================================================

@app.get("/api/dashboard/stats")
async def api_dashboard_stats(request: Request):
    try:
        stats = db.get_dashboard_stats()
        return JSONResponse(content=stats)
    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)


# ============================================================
# API — DEVICES
# ============================================================

@app.get("/api/devices")
async def api_get_devices(request: Request):
    try:
        devices = db.get_all_devices()
        return JSONResponse(content={"devices": devices})
    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)


@app.get("/api/devices/{ip_address}/uptime")
async def api_device_uptime(ip_address: str):
    """Returneaza istoricul uptime pentru un device."""
    try:
        ip     = ip_address.replace("-", ".")
        conn   = db.get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT TOP 24
                ul.checked_at, ul.is_online, ul.response_ms
            FROM uptime_log ul
            JOIN devices d ON ul.device_id = d.id
            WHERE d.ip_address = ?
            ORDER BY ul.checked_at DESC
        """, (ip,))
        columns = [col[0] for col in cursor.description]
        rows    = []
        for row in cursor.fetchall():
            item = dict(zip(columns, row))
            if item["checked_at"]:
                item["checked_at"] = item["checked_at"].strftime("%H:%M")
            item["is_online"] = bool(item["is_online"])
            rows.append(item)
        conn.close()
        return JSONResponse(content={"uptime": rows})
    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)


@app.get("/api/devices/{ip_address}")
async def api_get_device(ip_address: str):
    try:
        ip     = ip_address.replace("-", ".")
        device = db.get_device_by_ip(ip)
        if not device:
            return JSONResponse(content={"error": "Device negasit"}, status_code=404)
        return JSONResponse(content=device)
    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)


@app.put("/api/devices/{device_id}")
async def api_update_device(device_id: int, request: Request):
    try:
        data = await request.json()
        db.update_device_info(device_id, data)

        # Logam editarea
        user = auth_manager.get_current_user(request)
        if user:
            auth_manager.log_action(
                action      = "DEVICE_EDIT",
                user_id     = user["user_id"],
                username    = user["username"],
                entity_type = "device",
                entity_id   = device_id,
                details     = f"Owner: {data.get('owner')}, Dept: {data.get('department')}",
                ip_address  = request.client.host
            )

        return JSONResponse(content={"message": "Device actualizat cu succes!"})
    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)


# ============================================================
# API — SCAN
# ============================================================

@app.post("/api/scan")
async def api_start_scan(request: Request, background_tasks: BackgroundTasks):
    try:
        data     = await request.json()
        ip_range = data.get("ip_range", "").strip()
        detailed = data.get("detailed", False)

        if not ip_range:
            return JSONResponse(
                content={"error": "ip_range este obligatoriu"},
                status_code=400
            )

        if detailed:
            background_tasks.add_task(scanner.scan_with_details, ip_range)
        else:
            background_tasks.add_task(scanner.scan_network, ip_range)

        # Logam scanarea
        user = auth_manager.get_current_user(request)
        if user:
            auth_manager.log_action(
                action     = "SCAN_START",
                user_id    = user["user_id"],
                username   = user["username"],
                details    = f"Scanare {'detaliata' if detailed else 'rapida'}: {ip_range}",
                ip_address = request.client.host
            )

        return JSONResponse(content={
            "message":  f"Scanare pornita pentru {ip_range}",
            "ip_range": ip_range,
            "detailed": detailed
        })

    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)


@app.get("/api/scan/history")
async def api_scan_history():
    try:
        history = db.get_scan_history()
        return JSONResponse(content={"history": history})
    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)


# ============================================================
# API — ALERTS
# ============================================================

@app.get("/api/alerts")
async def api_get_alerts(unread_only: bool = False):
    try:
        alerts = db.get_alerts(unread_only=unread_only)
        return JSONResponse(content={"alerts": alerts})
    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)


@app.put("/api/alerts/{alert_id}/read")
async def api_mark_alert_read(alert_id: int):
    try:
        db.mark_alert_read(alert_id)
        return JSONResponse(content={"message": "Alerta marcata ca citita."})
    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)


# ============================================================
# API — AUDIT LOGS
# ============================================================

@app.get("/api/audit/logs")
async def api_audit_logs(request: Request):
    try:
        logs = auth_manager.get_audit_logs(limit=200)
        return JSONResponse(content={"logs": logs})
    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)


# ============================================================
# API — EXPORT
# ============================================================

@app.get("/api/export/excel")
async def export_excel(request: Request):
    try:
        devices = db.get_all_devices()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventar IT"

        header_fill  = PatternFill("solid", fgColor="1e3a5f")
        header_font  = Font(bold=True, color="FFFFFF", size=11)
        header_align = Alignment(horizontal="center", vertical="center")
        online_fill  = PatternFill("solid", fgColor="d1fae5")
        offline_fill = PatternFill("solid", fgColor="fee2e2")

        headers = [
            "ID", "IP Address", "Hostname", "Tip Device",
            "Sistem de Operare", "OS Family", "MAC Address",
            "Owner", "Departament", "Status",
            "Porturi Deschise", "Prima Descoperire",
            "Ultima Activitate", "Note"
        ]

        for col_num, header in enumerate(headers, 1):
            cell           = ws.cell(row=1, column=col_num, value=header)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align

        for row_num, device in enumerate(devices, 2):
            row_data = [
                device.get("id"), device.get("ip_address"),
                device.get("hostname"), device.get("device_type"),
                device.get("os_name"), device.get("os_family"),
                device.get("mac_address"), device.get("owner"),
                device.get("department"), device.get("status"),
                device.get("open_ports"), device.get("first_seen"),
                device.get("last_seen"), device.get("notes"),
            ]
            for col_num, value in enumerate(row_data, 1):
                cell           = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(vertical="center")

            status = device.get("status", "")
            if status == "online":
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_num, column=col).fill = online_fill
            elif status == "offline":
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_num, column=col).fill = offline_fill

        col_widths = [6, 16, 20, 14, 22, 12, 18, 18, 14, 10, 18, 20, 20, 30]
        for col_num, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

        ws.freeze_panes = "A2"

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Logam exportul
        user = auth_manager.get_current_user(request)
        if user:
            auth_manager.log_action(
                action     = "EXPORT_EXCEL",
                user_id    = user["user_id"],
                username   = user["username"],
                details    = f"Export Excel — {len(devices)} device-uri",
                ip_address = request.client.host
            )

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=inventar_DIIP.xlsx"}
        )

    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)


@app.get("/api/export/csv")
async def export_csv(request: Request):
    try:
        devices = db.get_all_devices()
        import io
        text_buffer = io.StringIO()
        headers     = [
            "id", "ip_address", "hostname", "device_type",
            "os_name", "os_family", "mac_address",
            "owner", "department", "status",
            "open_ports", "first_seen", "last_seen", "notes"
        ]
        writer = csv.DictWriter(text_buffer, fieldnames=headers, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(devices)
        buffer = BytesIO(text_buffer.getvalue().encode("utf-8-sig"))

        # Logam exportul
        user = auth_manager.get_current_user(request)
        if user:
            auth_manager.log_action(
                action     = "EXPORT_CSV",
                user_id    = user["user_id"],
                username   = user["username"],
                details    = f"Export CSV — {len(devices)} device-uri",
                ip_address = request.client.host
            )

        return StreamingResponse(
            buffer,
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=inventar_DIIP.csv"}
        )

    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)


# ============================================================
# API — EMAIL TEST
# ============================================================

@app.get("/api/email/test")
async def api_test_email():
    result = email_service.test_email()
    return JSONResponse(content=result)


# ============================================================
# API — HEALTH CHECK
# ============================================================

@app.get("/api/health")
async def api_health():
    db_status = db.test_connection()
    return JSONResponse(content={
        "app":      "running",
        "database": db_status["status"],
        "message":  db_status.get("message")
    })


# ============================================================
# PORNIRE DIRECTA
# ============================================================
if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8080, reload=True)


# ============================================================
# ANALYTICS ROUTES — adauga in main.py
# ============================================================

@app.get("/analytics")
async def page_analytics(request: Request):
    user = auth_manager.get_current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=302)
    return templates.TemplateResponse(
        request=request, name="analytics.html", context={"user": user}
    )


@app.get("/api/analytics/kpi")
async def api_analytics_kpi(days: int = 30):
    """KPI-uri principale pentru perioada selectata."""
    try:
        conn = db.get_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT
                ROUND(AVG(CAST(pct AS FLOAT)), 1) AS avg_uptime
            FROM (
                SELECT
                    device_id,
                    ROUND(
                        SUM(CASE WHEN is_online = 1 THEN 1.0 ELSE 0 END)
                        / NULLIF(COUNT(*), 0) * 100, 1
                    ) AS pct
                FROM uptime_log
                WHERE checked_at >= DATEADD(day, -?, GETDATE())
                GROUP BY device_id
            ) sub
        """, (days,))
        row = cursor.fetchone()
        avg_uptime = round(row[0], 1) if row and row[0] else 0

        cursor.execute("SELECT COUNT(*) FROM devices")
        total_devices = cursor.fetchone()[0]

        cursor.execute(
            "SELECT COUNT(*) FROM alerts WHERE created_at >= DATEADD(day, -?, GETDATE())",
            (days,)
        )
        total_alerts = cursor.fetchone()[0]

        cursor.execute(
            "SELECT COUNT(*) FROM scan_history WHERE started_at >= DATEADD(day, -?, GETDATE())",
            (days,)
        )
        total_scans = cursor.fetchone()[0]

        conn.close()
        return JSONResponse(content={
            "avg_uptime": avg_uptime,
            "total_devices": total_devices,
            "total_alerts": total_alerts,
            "total_scans": total_scans
        })
    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)


@app.get("/api/analytics/activity")
async def api_analytics_activity(days: int = 30):
    """Activitate zilnica: scanari + alerte."""
    try:
        conn = db.get_connection()
        cursor = conn.cursor()

        cursor.execute("""
            WITH activity AS (
                SELECT CAST(started_at AS DATE) AS zi, 'scanare' AS tip, COUNT(*) AS numar
                FROM scan_history
                WHERE started_at >= DATEADD(day, -?, GETDATE())
                GROUP BY CAST(started_at AS DATE)
                UNION ALL
                SELECT CAST(created_at AS DATE) AS zi, 'alerta' AS tip, COUNT(*) AS numar
                FROM alerts
                WHERE created_at >= DATEADD(day, -?, GETDATE())
                GROUP BY CAST(created_at AS DATE)
            )
            SELECT
                zi,
                SUM(CASE WHEN tip = 'scanare' THEN numar ELSE 0 END) AS scanari,
                SUM(CASE WHEN tip = 'alerta'  THEN numar ELSE 0 END) AS alerte
            FROM activity
            GROUP BY zi
            ORDER BY zi DESC
        """, (days, days))

        columns = [col[0] for col in cursor.description]
        rows = []
        for row in cursor.fetchall():
            item = dict(zip(columns, row))
            if item.get("zi"):
                item["zi"] = str(item["zi"])
            rows.append(item)

        conn.close()
        return JSONResponse(content={"activity": rows})
    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)


@app.get("/api/analytics/departments")
async def api_analytics_departments():
    """Distributia device-urilor pe departament cu procente."""
    try:
        conn = db.get_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT
                ISNULL(department, 'Neatribuit') AS departament,
                COUNT(*) AS total_devices,
                SUM(CASE WHEN status = 'online'  THEN 1 ELSE 0 END) AS online,
                SUM(CASE WHEN status = 'offline' THEN 1 ELSE 0 END) AS offline,
                ROUND(COUNT(*) * 100.0 / SUM(COUNT(*)) OVER(), 1)   AS procent
            FROM devices
            GROUP BY department
            ORDER BY total_devices DESC
        """)

        columns = [col[0] for col in cursor.description]
        rows = [dict(zip(columns, row)) for row in cursor.fetchall()]
        conn.close()
        return JSONResponse(content={"departments": rows})
    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)


@app.get("/api/analytics/uptime-ranking")
async def api_analytics_uptime_ranking():
    """Top device-uri dupa uptime cu RANK() si NTILE()."""
    try:
        conn = db.get_connection()
        cursor = conn.cursor()

        cursor.execute("""
            WITH uptime_stats AS (
                SELECT
                    d.id, d.ip_address, d.hostname, d.device_type, d.department,
                    COUNT(ul.id) AS total_checks,
                    ROUND(
                        SUM(CASE WHEN ul.is_online = 1 THEN 1.0 ELSE 0 END)
                        / NULLIF(COUNT(ul.id), 0) * 100, 1
                    ) AS uptime_pct,
                    AVG(CASE WHEN ul.response_ms IS NOT NULL
                        THEN CAST(ul.response_ms AS FLOAT) END) AS avg_response_ms
                FROM devices d
                LEFT JOIN uptime_log ul ON d.id = ul.device_id
                GROUP BY d.id, d.ip_address, d.hostname, d.device_type, d.department
            )
            SELECT
                ip_address,
                ISNULL(hostname, 'Unknown') AS hostname,
                device_type,
                ISNULL(department, 'Neatribuit') AS department,
                total_checks,
                ISNULL(uptime_pct, 0) AS uptime_pct,
                ROUND(avg_response_ms, 0) AS avg_ms,
                RANK()  OVER (ORDER BY uptime_pct DESC) AS rank_uptime,
                NTILE(4) OVER (ORDER BY uptime_pct DESC) AS quartile
            FROM uptime_stats
            WHERE total_checks > 0
            ORDER BY rank_uptime
        """)

        columns = [col[0] for col in cursor.description]
        rows = []
        for row in cursor.fetchall():
            item = dict(zip(columns, row))
            for k, v in item.items():
                if hasattr(v, 'strftime'):
                    item[k] = str(v)
            rows.append(item)

        conn.close()
        return JSONResponse(content={"ranking": rows})
    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)


@app.get("/api/analytics/alerts-analysis")
async def api_analytics_alerts_analysis():
    """Analiza alertelor pe tip si severitate."""
    try:
        conn = db.get_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT
                alert_type AS tip_alerta,
                COUNT(*) AS total,
                SUM(CASE WHEN severity = 'critical' THEN 1 ELSE 0 END) AS critice,
                SUM(CASE WHEN severity = 'warning'  THEN 1 ELSE 0 END) AS avertismente,
                SUM(CASE WHEN severity = 'info'     THEN 1 ELSE 0 END) AS informatii,
                SUM(CASE WHEN is_read = 0           THEN 1 ELSE 0 END) AS necitite
            FROM alerts
            GROUP BY alert_type
            ORDER BY total DESC
        """)

        columns = [col[0] for col in cursor.description]
        rows = [dict(zip(columns, row)) for row in cursor.fetchall()]
        conn.close()
        return JSONResponse(content={"analysis": rows})
    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)


@app.get("/api/analytics/evolution")
async def api_analytics_evolution():
    """Evolutia cumulativa a device-urilor descoperite."""
    try:
        conn = db.get_connection()
        cursor = conn.cursor()

        cursor.execute("""
            WITH device_timeline AS (
                SELECT
                    CAST(first_seen AS DATE) AS data_descoperire,
                    COUNT(*) AS device_noi,
                    SUM(COUNT(*)) OVER (
                        ORDER BY CAST(first_seen AS DATE)
                        ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW
                    ) AS total_cumulativ
                FROM devices
                WHERE first_seen IS NOT NULL
                GROUP BY CAST(first_seen AS DATE)
            )
            SELECT
                data_descoperire,
                device_noi,
                total_cumulativ
            FROM device_timeline
            ORDER BY data_descoperire DESC
        """)

        columns = [col[0] for col in cursor.description]
        rows = []
        for row in cursor.fetchall():
            item = dict(zip(columns, row))
            if item.get("data_descoperire"):
                item["data_descoperire"] = str(item["data_descoperire"])
            rows.append(item)

        conn.close()
        return JSONResponse(content={"evolution": rows})
    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)
